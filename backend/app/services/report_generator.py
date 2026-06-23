from __future__ import annotations

import sys
from html import escape
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
from sqlalchemy import select, text
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models import DwdAgriAkpDensityDaily, DwdAgriAttendanceDaily, UploadBatch


HARVEST_REPORT_NAME = "产量监控日报"
REPO_ROOT = Path(__file__).resolve().parents[3]
SCRIPTS_DIR = REPO_ROOT / "scripts"


def generate_batch_report(batch: UploadBatch, db: Session) -> str:
    return generate_harvest_summary_report(batch, db)


def generate_harvest_summary_report(batch: UploadBatch, db: Session) -> str:
    refresh_harvest_data_mart()
    rows = fetch_harvest_summary_rows(batch, db)
    report_date = rows[0]["report_date"] if rows else batch.report_date
    display_batch = SimpleNamespace(
        batch_no=batch.batch_no,
        report_date=report_date,
        department=batch.department,
        site=batch.site,
    )
    report_paths = get_generated_report_paths(batch)
    workbook_paths = get_generated_report_workbook_paths(batch)
    image_paths = get_generated_report_image_paths(batch)
    output_paths = [
        write_report(report_paths[HARVEST_REPORT_NAME], build_harvest_summary_page(display_batch, rows)),
        write_harvest_summary_workbook(workbook_paths[HARVEST_REPORT_NAME], display_batch, rows),
        write_harvest_summary_image(image_paths[HARVEST_REPORT_NAME], display_batch, rows),
    ]
    batch.status = "report_generated"
    return "; ".join(str(path) for path in output_paths)


def refresh_harvest_data_mart() -> None:
    if str(SCRIPTS_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPTS_DIR))

    import psycopg
    from build_harvest_dwd import build_dwd
    from build_harvest_production_summary import build_summary

    dsn = settings.database_url.replace("postgresql+psycopg://", "postgresql://", 1)
    with psycopg.connect(dsn) as connection:
        build_dwd(connection)
        build_summary(connection)


def fetch_harvest_summary_rows(batch: UploadBatch, db: Session) -> list[dict[str, Any]]:
    report_date = resolve_harvest_report_date(batch, db)
    result = db.execute(
        text(
            """
            select
                report_date,
                coalesce(estate_name, report_site_name, estate_code, '') as estate_name,
                division_code,
                mature_area_ha,
                bbc_estimated_production_kg,
                akp_density_percent,
                akp_estimated_harvest_area_ha,
                akp_estimated_production_kg,
                daily_actual_production_kg,
                daily_production_completion_rate,
                month_to_date_actual_production_kg,
                monthly_bbc_target_kg,
                bbc_target_completion_rate,
                bbc_remaining_target_kg,
                monthly_budget_production_kg,
                budget_completion_rate,
                budget_remaining_target_kg,
                month_remaining_days,
                data_note
            from dws.harvest_production_daily_summary
            where report_date = :report_date
            order by estate_name, division_code
            """
        ),
        {"report_date": report_date},
    )
    return [dict(row._mapping) for row in result]


def resolve_harvest_report_date(batch: UploadBatch, db: Session) -> Any:
    result = db.execute(
        text(
            """
            select report_date
            from dws.harvest_production_daily_summary
            where report_date <= :report_date
              and daily_actual_production_kg is not null
            group by report_date
            having sum(coalesce(daily_actual_production_kg, 0)) > 0
            order by report_date desc
            limit 1
            """
        ),
        {"report_date": batch.report_date},
    ).first()
    return result[0] if result else batch.report_date


def _generate_legacy_batch_report(batch: UploadBatch, db: Session) -> str:
    report_paths = get_generated_report_paths(batch)

    akp_rows = list(
        db.scalars(
            select(DwdAgriAkpDensityDaily)
            .where(DwdAgriAkpDensityDaily.batch_id == batch.id)
            .order_by(DwdAgriAkpDensityDaily.id)
        ).all()
    )
    attendance_rows = list(
        db.scalars(
            select(DwdAgriAttendanceDaily)
            .where(
                DwdAgriAttendanceDaily.batch_id == batch.id,
                DwdAgriAttendanceDaily.section == "daily_attendance",
            )
            .order_by(DwdAgriAttendanceDaily.id)
        ).all()
    )

    workbook_paths = get_generated_report_workbook_paths(batch)
    image_paths = get_generated_report_image_paths(batch)
    output_paths = [
        write_report(
            report_paths["铲果密度"],
            build_page(
                batch=batch,
                title="铲果密度",
                subtitle="DWD 标准明细",
                body=build_akp_body(akp_rows),
            ),
        ),
        write_akp_workbook(
            workbook_paths["铲果密度"],
            batch,
            akp_rows,
        ),
        write_akp_image(
            image_paths["铲果密度"],
            batch,
            akp_rows,
        ),
        write_report(
            report_paths["铲果工养护工出勤"],
            build_page(
                batch=batch,
                title="铲果工养护工出勤",
                subtitle="DWD 标准明细",
                body=build_attendance_body(attendance_rows),
            ),
        ),
        write_attendance_workbook(
            workbook_paths["铲果工养护工出勤"],
            batch,
            attendance_rows,
        ),
        write_attendance_image(
            image_paths["铲果工养护工出勤"],
            batch,
            attendance_rows,
        ),
    ]

    batch.status = "report_generated"
    return "; ".join(str(path) for path in output_paths)


def _report_file_stem(batch: UploadBatch, report_name: str) -> str:
    """固定文件名：日期_报表名_园区，相同三要素的报表生成时直接覆盖旧文件。"""
    parts = [batch.report_date.isoformat(), report_name]
    if batch.site:
        parts.append(batch.site)
    return "_".join(parts)


def get_generated_report_paths(batch: UploadBatch) -> dict[str, Path]:
    report_date = batch.report_date.isoformat()
    date_parts = report_date.split("-")
    report_root = Path(settings.storage_dir) / "reports" / date_parts[0] / date_parts[1] / date_parts[2]
    stem = _report_file_stem(batch, HARVEST_REPORT_NAME)
    return {
        HARVEST_REPORT_NAME: report_root / HARVEST_REPORT_NAME / f"{stem}.html",
    }


def get_generated_report_workbook_paths(batch: UploadBatch) -> dict[str, Path]:
    report_date = batch.report_date.isoformat()
    date_parts = report_date.split("-")
    report_root = Path(settings.storage_dir) / "reports" / date_parts[0] / date_parts[1] / date_parts[2]
    stem = _report_file_stem(batch, HARVEST_REPORT_NAME)
    return {
        HARVEST_REPORT_NAME: report_root / HARVEST_REPORT_NAME / f"{stem}.xlsx",
    }


def get_generated_report_image_paths(batch: UploadBatch) -> dict[str, Path]:
    report_date = batch.report_date.isoformat()
    date_parts = report_date.split("-")
    report_root = Path(settings.storage_dir) / "reports" / date_parts[0] / date_parts[1] / date_parts[2]
    stem = _report_file_stem(batch, HARVEST_REPORT_NAME)
    return {
        HARVEST_REPORT_NAME: report_root / HARVEST_REPORT_NAME / f"{stem}.png",
    }


def build_harvest_summary_page(batch: UploadBatch, rows: list[dict[str, Any]]) -> str:
    body = build_harvest_summary_body(rows) if rows else '<div class="empty">没有找到该日期的产量监控汇总数据。</div>'
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <title>{h(str(batch.report_date))}_{h(HARVEST_REPORT_NAME)}</title>
  <style>
    body {{ margin: 0; padding: 28px; background: #f4f7fb; color: #172033; font-family: "Microsoft YaHei", Arial, sans-serif; }}
    .report {{ max-width: 1680px; margin: 0 auto; background: #fff; border: 1px solid #dbe5f0; border-radius: 8px; overflow: hidden; }}
    header {{ padding: 26px 30px; border-bottom: 1px solid #e2e8f0; background: #f8fafc; }}
    h1 {{ margin: 0 0 8px; font-size: 28px; }}
    .meta {{ display: flex; gap: 24px; color: #475569; }}
    section {{ padding: 24px 30px 30px; }}
    .summary {{ display: grid; grid-template-columns: repeat(5, minmax(0, 1fr)); gap: 12px; margin-bottom: 18px; }}
    .summary div {{ padding: 12px 14px; border: 1px solid #dbe5f0; border-radius: 8px; background: #f8fafc; }}
    .summary span {{ display: block; color: #64748b; font-size: 12px; }}
    .summary strong {{ display: block; margin-top: 5px; font-size: 20px; }}
    .table-wrap {{ overflow-x: auto; border: 1px solid #dbe5f0; border-radius: 8px; }}
    table {{ width: 100%; border-collapse: collapse; min-width: 1500px; font-size: 13px; }}
    th, td {{ padding: 9px 10px; border-bottom: 1px solid #e2e8f0; text-align: left; white-space: nowrap; }}
    th {{ background: #eef6f5; color: #0f766e; font-weight: 700; }}
    .empty {{ padding: 18px; border: 1px dashed #cbd5e1; border-radius: 8px; color: #64748b; background: #f8fafc; }}
  </style>
</head>
<body>
  <article class="report">
    <header>
      <h1>{h(str(batch.report_date))}_{h(HARVEST_REPORT_NAME)}</h1>
      <div class="meta">
        <span>批次号：{h(batch.batch_no)}</span>
        <span>部门：{h(batch.department)}</span>
        <span>园区/工厂：{h(batch.site or "-")}</span>
      </div>
    </header>
    <section>{body}</section>
  </article>
</body>
</html>"""


def build_harvest_summary_body(rows: list[dict[str, Any]]) -> str:
    table_rows = "\n".join(
        f"""<tr>
          <td>{h(str(row.get("estate_name") or ""))}</td>
          <td>{h(str(row.get("division_code") or ""))}</td>
          <td>{fmt(row.get("mature_area_ha"))}</td>
          <td>{fmt(row.get("bbc_estimated_production_kg"))}</td>
          <td>{fmt(row.get("akp_density_percent"))}%</td>
          <td>{fmt(row.get("akp_estimated_harvest_area_ha"))}</td>
          <td>{fmt(row.get("akp_estimated_production_kg"))}</td>
          <td>{fmt(row.get("daily_actual_production_kg"))}</td>
          <td>{fmt_percent(row.get("daily_production_completion_rate"))}</td>
          <td>{fmt(row.get("month_to_date_actual_production_kg"))}</td>
          <td>{fmt(row.get("monthly_bbc_target_kg"))}</td>
          <td>{fmt_percent(row.get("bbc_target_completion_rate"))}</td>
          <td>{fmt(row.get("bbc_remaining_target_kg"))}</td>
          <td>{fmt(row.get("monthly_budget_production_kg"))}</td>
          <td>{fmt_percent(row.get("budget_completion_rate"))}</td>
          <td>{fmt(row.get("budget_remaining_target_kg"))}</td>
          <td>{fmt(row.get("month_remaining_days"))}</td>
          <td>{h(str(row.get("data_note") or ""))}</td>
        </tr>"""
        for row in rows
    )
    daily_actual = sum_decimal(row.get("daily_actual_production_kg") for row in rows)
    mtd_actual = sum_decimal(row.get("month_to_date_actual_production_kg") for row in rows)
    bbc_target = sum_decimal(row.get("monthly_bbc_target_kg") for row in rows)
    month_target = sum_decimal(row.get("monthly_budget_production_kg") for row in rows)
    return f"""
      <div class="summary">
        <div><span>当日总产量</span><strong>{fmt(daily_actual)}</strong></div>
        <div><span>BBC目标</span><strong>{fmt(bbc_target)}</strong></div>
        <div><span>BBC完成率</span><strong>{fmt_percent(safe_ratio(mtd_actual, bbc_target))}</strong></div>
        <div><span>月目标</span><strong>{fmt(month_target)}</strong></div>
        <div><span>月目标完成率</span><strong>{fmt_percent(safe_ratio(mtd_actual, month_target))}</strong></div>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>公司</th><th>小区</th><th>成熟面积</th><th>BBC</th><th>AKP密度</th>
              <th>AKP预计铲果面积</th><th>AKP预计产量</th><th>当日实际产量</th><th>当日产量完成率</th>
              <th>月累计实际产量</th><th>月度BBC目标产量</th><th>BBC目标完成率</th><th>BBC剩余目标产量</th>
              <th>月目标产量</th><th>月目标完成率</th><th>月目标剩余产量</th><th>月剩余天数</th><th>数据说明</th>
            </tr>
          </thead>
          <tbody>{table_rows}</tbody>
        </table>
      </div>"""


def write_harvest_summary_workbook(path: Path, batch: UploadBatch, rows: list[dict[str, Any]]) -> Path:
    headers = [
        "公司", "小区", "成熟面积", "BBC", "AKP密度", "AKP预计铲果面积", "AKP预计产量",
        "当日实际产量", "当日产量完成率", "月累计实际产量", "月度BBC目标产量", "BBC目标完成率",
        "BBC剩余目标产量", "月目标产量", "月目标完成率", "月目标剩余产量", "月剩余天数", "数据说明",
    ]
    data_rows = [
        [
            row.get("estate_name"),
            row.get("division_code"),
            row.get("mature_area_ha"),
            row.get("bbc_estimated_production_kg"),
            row.get("akp_density_percent"),
            row.get("akp_estimated_harvest_area_ha"),
            row.get("akp_estimated_production_kg"),
            row.get("daily_actual_production_kg"),
            row.get("daily_production_completion_rate"),
            row.get("month_to_date_actual_production_kg"),
            row.get("monthly_bbc_target_kg"),
            row.get("bbc_target_completion_rate"),
            row.get("bbc_remaining_target_kg"),
            row.get("monthly_budget_production_kg"),
            row.get("budget_completion_rate"),
            row.get("budget_remaining_target_kg"),
            row.get("month_remaining_days"),
            row.get("data_note"),
        ]
        for row in rows
    ]
    return write_workbook(path, batch, HARVEST_REPORT_NAME, headers, data_rows, percent_columns={9, 12, 15})


def write_harvest_summary_image(path: Path, batch: UploadBatch, rows: list[dict[str, Any]]) -> Path:
    daily_actual = sum_decimal(row.get("daily_actual_production_kg") for row in rows)
    mtd_actual = sum_decimal(row.get("month_to_date_actual_production_kg") for row in rows)
    bbc_target = sum_decimal(row.get("monthly_bbc_target_kg") for row in rows)
    month_target = sum_decimal(row.get("monthly_budget_production_kg") for row in rows)
    summary = [
        ("当日总产量", fmt(daily_actual)),
        ("BBC目标", fmt(bbc_target)),
        ("BBC完成率", fmt_percent(safe_ratio(mtd_actual, bbc_target))),
        ("月目标", fmt(month_target)),
        ("月目标完成率", fmt_percent(safe_ratio(mtd_actual, month_target))),
    ]
    columns = [
        ("公司", 120),
        ("小区", 80),
        ("成熟面积", 120),
        ("AKP预计产量", 135),
        ("实际产量", 120),
        ("完成率", 100),
        ("月累计", 130),
        ("BBC目标", 120),
        ("BBC完成率", 120),
        ("月目标", 120),
        ("月目标完成率", 120),
        ("剩余天数", 100),
        ("数据说明", 340),
    ]
    table_rows = [
        [
            row.get("estate_name"),
            row.get("division_code"),
            fmt(row.get("mature_area_ha")),
            fmt(row.get("akp_estimated_production_kg")),
            fmt(row.get("daily_actual_production_kg")),
            fmt_percent(row.get("daily_production_completion_rate")),
            fmt(row.get("month_to_date_actual_production_kg")),
            fmt(row.get("monthly_bbc_target_kg")),
            fmt_percent(row.get("bbc_target_completion_rate")),
            fmt(row.get("monthly_budget_production_kg")),
            fmt_percent(row.get("budget_completion_rate")),
            fmt(row.get("month_remaining_days")),
            row.get("data_note"),
        ]
        for row in rows
    ]
    return write_table_image(path, batch, HARVEST_REPORT_NAME, summary, columns, table_rows)


def write_report(path: Path, html: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html, encoding="utf-8")
    return path


def write_akp_image(path: Path, batch: UploadBatch, rows: list[DwdAgriAkpDensityDaily]) -> Path:
    detail_rows = [row for row in rows if row.row_label == "detail"]
    summary = [
        ("明细行数", str(len(detail_rows))),
        ("总 Panen Kg", fmt(sum_number(row.panen_kg for row in detail_rows))),
        ("总 Janjang", fmt(sum_number(row.jumlah_janjang for row in detail_rows))),
    ]
    columns = [
        ("小区", 90),
        ("Blok", 150),
        ("SAP", 90),
        ("Luas Ha", 120),
        ("TT", 80),
        ("Panen", 110),
        ("AKP", 90),
        ("Panen Kg", 135),
        ("Janjang", 130),
        ("TK Panen", 120),
        ("Keterangan", 170),
        ("行类型", 110),
    ]
    table_rows = [
        [
            row.division,
            row.blok,
            row.sap,
            fmt(row.luas_ha),
            fmt(row.tt_year),
            fmt(row.panen_count),
            f"{fmt(row.akp_percent)}%" if row.akp_percent is not None else "",
            fmt(row.panen_kg),
            fmt(row.jumlah_janjang),
            fmt(row.tk_panen),
            row.keterangan,
            row.row_label,
        ]
        for row in rows
    ]
    return write_table_image(path, batch, "铲果密度", summary, columns, table_rows)


def write_attendance_image(path: Path, batch: UploadBatch, rows: list[DwdAgriAttendanceDaily]) -> Path:
    detail_rows = [row for row in rows if row.row_label == "detail"]
    harvester_rows = [row for row in detail_rows if row.worker_type == "harvester"]
    maintenance_rows = [row for row in detail_rows if row.worker_type == "maintenance"]
    summary = [
        ("铲果工出勤率", average_percent(row.hadir_percent for row in harvester_rows)),
        ("铲果工出勤人数", fmt(sum_number(row.hadir for row in harvester_rows))),
        ("养护工出勤率", average_percent(row.hadir_percent for row in maintenance_rows)),
        ("养护工出勤人数", fmt(sum_number(row.hadir for row in maintenance_rows))),
    ]
    columns = [
        ("人员类型", 130),
        ("Afdeling", 120),
        ("Luas Ha", 120),
        ("需求人数", 120),
        ("实际人数", 120),
        ("差异", 90),
        ("出勤人数", 120),
        ("出勤率", 100),
        ("Ijin", 90),
        ("Ijin率", 100),
        ("总人数", 110),
        ("总率", 100),
        ("行类型", 110),
    ]
    table_rows = [
        [
            worker_type_label(row.worker_type),
            row.afdeling,
            fmt(row.luas_ha),
            fmt(row.kebutuhan_pemanen),
            fmt(row.actual_pemanen),
            fmt(row.actual_vs_kebutuhan),
            fmt(row.hadir),
            f"{fmt(row.hadir_percent)}%" if row.hadir_percent is not None else "",
            fmt(row.ijin),
            f"{fmt(row.ijin_percent)}%" if row.ijin_percent is not None else "",
            fmt(row.total_karyawan),
            f"{fmt(row.total_percent)}%" if row.total_percent is not None else "",
            row.row_label,
        ]
        for row in rows
    ]
    return write_table_image(path, batch, "铲果工养护工出勤", summary, columns, table_rows)


def write_table_image(
    path: Path,
    batch: UploadBatch,
    title: str,
    summary: list[tuple[str, str]],
    columns: list[tuple[str, int]],
    rows: list[list[Any]],
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fonts = load_image_fonts()
    margin = 48
    table_width = sum(width for _, width in columns)
    image_width = table_width + margin * 2
    title_height = 160
    summary_height = 110 if summary else 0
    row_height = 50
    header_height = 54
    image_height = margin + title_height + summary_height + header_height + max(len(rows), 1) * row_height + margin

    image = Image.new("RGB", (image_width, image_height), "#f4f7fb")
    draw = ImageDraw.Draw(image)
    draw.rounded_rectangle(
        (24, 24, image_width - 24, image_height - 24),
        radius=18,
        fill="#ffffff",
        outline="#dbe5f0",
        width=2,
    )
    y = margin
    draw.text((margin, y), f"{batch.report_date}_{title}", font=fonts["title"], fill="#172033")
    y += 62
    meta = f"批次号：{batch.batch_no}    部门：{batch.department}    园区/工厂：{batch.site or '-'}"
    draw.text((margin, y), meta, font=fonts["meta"], fill="#475569")
    y += 68

    if summary:
        box_gap = 14
        box_width = (table_width - box_gap * (len(summary) - 1)) // len(summary)
        for index, (label, value) in enumerate(summary):
            x = margin + index * (box_width + box_gap)
            draw.rounded_rectangle((x, y, x + box_width, y + 84), radius=12, fill="#f8fafc", outline="#dbe5f0")
            draw.text((x + 18, y + 14), label, font=fonts["small"], fill="#64748b")
            draw.text((x + 18, y + 42), value, font=fonts["summary"], fill="#172033")
        y += 110

    x = margin
    draw.rectangle((margin, y, margin + table_width, y + header_height), fill="#eef6f5")
    for label, width in columns:
        draw_cell_text(draw, (x + 10, y, x + width - 10, y + header_height), label, fonts["header"], "#0f766e")
        draw.line((x, y, x, y + header_height + max(len(rows), 1) * row_height), fill="#dbe5f0", width=1)
        x += width
    draw.line((margin + table_width, y, margin + table_width, y + header_height + max(len(rows), 1) * row_height), fill="#dbe5f0", width=1)
    y += header_height

    if not rows:
        draw.text((margin + 18, y + 16), "没有可展示的数据。", font=fonts["body"], fill="#64748b")
    for row_index, row in enumerate(rows):
        fill = "#ffffff" if row_index % 2 == 0 else "#f8fafc"
        if str(row[-1]).lower() in {"subtotal", "total"}:
            fill = "#ecfdf5"
        draw.rectangle((margin, y, margin + table_width, y + row_height), fill=fill)
        x = margin
        for value, (_, width) in zip(row, columns):
            draw_cell_text(draw, (x + 10, y, x + width - 10, y + row_height), str(value or ""), fonts["body"], "#172033")
            x += width
        draw.line((margin, y + row_height, margin + table_width, y + row_height), fill="#e2e8f0", width=1)
        y += row_height

    image.save(path, format="PNG", optimize=True)
    return path


def load_image_fonts() -> dict[str, ImageFont.FreeTypeFont | ImageFont.ImageFont]:
    font_path = find_chinese_font_path()

    def font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
        if font_path:
            return ImageFont.truetype(str(font_path), size=size)
        return ImageFont.load_default()

    return {
        "title": font(38),
        "summary": font(30),
        "header": font(22),
        "body": font(20),
        "meta": font(22),
        "small": font(18),
    }


def find_chinese_font_path() -> Path | None:
    candidates = [
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/Alibaba-PuHuiTi-Regular.otf"),
        Path("C:/Windows/Fonts/simsun.ttc"),
    ]
    return next((path for path in candidates if path.exists()), None)


def draw_cell_text(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    text: str,
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    fill: str,
) -> None:
    left, top, right, bottom = box
    available_width = right - left
    clipped = clip_text(draw, text, font, available_width)
    text_box = draw.textbbox((0, 0), clipped, font=font)
    text_height = text_box[3] - text_box[1]
    y = top + max((bottom - top - text_height) // 2, 0)
    draw.text((left, y), clipped, font=font, fill=fill)


def clip_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont | ImageFont.ImageFont, width: int) -> str:
    if draw.textlength(text, font=font) <= width:
        return text
    ellipsis = "..."
    while text and draw.textlength(text + ellipsis, font=font) > width:
        text = text[:-1]
    return text + ellipsis if text else ellipsis


def write_akp_workbook(path: Path, batch: UploadBatch, rows: list[DwdAgriAkpDensityDaily]) -> Path:
    headers = [
        "小区",
        "Blok",
        "SAP",
        "Luas Ha",
        "TT",
        "Panen",
        "AKP%",
        "Panen Kg",
        "Janjang",
        "TK Panen",
        "Keterangan",
        "行类型",
    ]
    data_rows = [
        [
            row.division,
            row.blok,
            row.sap,
            row.luas_ha,
            row.tt_year,
            row.panen_count,
            percent_value(row.akp_percent),
            row.panen_kg,
            row.jumlah_janjang,
            row.tk_panen,
            row.keterangan,
            row.row_label,
        ]
        for row in rows
    ]
    return write_workbook(path, batch, "铲果密度", headers, data_rows, percent_columns={7})


def write_attendance_workbook(path: Path, batch: UploadBatch, rows: list[DwdAgriAttendanceDaily]) -> Path:
    headers = [
        "人员类型",
        "Afdeling",
        "Luas Ha",
        "需求人数",
        "实际人数",
        "差异",
        "出勤人数",
        "出勤率",
        "Ijin",
        "Ijin率",
        "总人数",
        "总率",
        "行类型",
    ]
    data_rows = [
        [
            worker_type_label(row.worker_type),
            row.afdeling,
            row.luas_ha,
            row.kebutuhan_pemanen,
            row.actual_pemanen,
            row.actual_vs_kebutuhan,
            row.hadir,
            percent_value(row.hadir_percent),
            row.ijin,
            percent_value(row.ijin_percent),
            row.total_karyawan,
            percent_value(row.total_percent),
            row.row_label,
        ]
        for row in rows
    ]
    return write_workbook(path, batch, "铲果工养护工出勤", headers, data_rows, percent_columns={8, 10, 12})


def write_workbook(
    path: Path,
    batch: UploadBatch,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    percent_columns: set[int],
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    sheet.freeze_panes = "A6"

    sheet["A1"] = f"{batch.report_date}_{title}"
    sheet["A1"].font = Font(bold=True, size=16, color="172033")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    meta = [
        ("批次号", batch.batch_no),
        ("部门", batch.department),
        ("园区/工厂", batch.site or "-"),
        ("报表日期", str(batch.report_date)),
    ]
    for column, (label, value) in enumerate(meta, start=1):
        sheet.cell(row=3, column=column, value=label).font = Font(bold=True, color="475569")
        sheet.cell(row=4, column=column, value=value)

    header_fill = PatternFill("solid", fgColor="EEF6F5")
    header_font = Font(bold=True, color="0F766E")
    border = Border(bottom=Side(style="thin", color="DBE5F0"))
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=6, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=7):
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.border = border
            if column in percent_columns and value is not None:
                cell.number_format = "0%"

    for column in range(1, len(headers) + 1):
        values = [sheet.cell(row=row, column=column).value for row in range(1, sheet.max_row + 1)]
        width = min(max(len(str(value)) if value is not None else 0 for value in values) + 3, 36)
        sheet.column_dimensions[get_column_letter(column)].width = max(width, 10)

    workbook.save(path)
    return path


def build_page(batch: UploadBatch, title: str, subtitle: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <title>{h(str(batch.report_date))}_{h(title)}</title>
  <style>
    :root {{
      color: #172033;
      background: #f4f7fb;
      font-family: "Microsoft YaHei", "PingFang SC", Arial, sans-serif;
    }}
    body {{ margin: 0; padding: 28px; }}
    .report {{
      max-width: 1440px;
      margin: 0 auto;
      background: #fff;
      border: 1px solid #dbe5f0;
      border-radius: 8px;
      box-shadow: 0 18px 45px rgba(15, 23, 42, 0.08);
      overflow: hidden;
    }}
    header {{
      padding: 26px 30px;
      border-bottom: 1px solid #e2e8f0;
      background: #f8fafc;
    }}
    h1 {{ margin: 0 0 6px; font-size: 26px; line-height: 1.3; }}
    .subtitle {{ margin: 0 0 16px; color: #64748b; }}
    .meta {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 12px;
      color: #475569;
      font-size: 14px;
    }}
    .meta strong {{ display: block; color: #0f172a; margin-top: 4px; }}
    section {{ padding: 24px 30px 30px; }}
    .summary {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 12px;
      margin-bottom: 18px;
    }}
    .summary div {{
      padding: 12px 14px;
      border: 1px solid #dbe5f0;
      border-radius: 8px;
      background: #f8fafc;
    }}
    .summary span {{ display: block; color: #64748b; font-size: 12px; }}
    .summary strong {{ display: block; margin-top: 5px; font-size: 20px; }}
    .table-wrap {{
      overflow-x: auto;
      border: 1px solid #dbe5f0;
      border-radius: 8px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 1000px;
      font-size: 13px;
    }}
    th, td {{
      padding: 9px 10px;
      border-bottom: 1px solid #e2e8f0;
      text-align: left;
      white-space: nowrap;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #eef6f5;
      color: #0f766e;
      font-weight: 700;
    }}
    tr.subtotal td, tr.total td {{ background: #f8fafc; font-weight: 700; }}
    .empty {{
      padding: 18px;
      border: 1px dashed #cbd5e1;
      border-radius: 8px;
      color: #64748b;
      background: #f8fafc;
    }}
  </style>
</head>
<body>
  <article class="report">
    <header>
      <h1>{h(str(batch.report_date))}_{h(title)}</h1>
      <p class="subtitle">{h(subtitle)}</p>
      <div class="meta">
        <div>批次号<strong>{h(batch.batch_no)}</strong></div>
        <div>部门<strong>{h(batch.department)}</strong></div>
        <div>园区/工厂<strong>{h(batch.site or "-")}</strong></div>
        <div>报表日期<strong>{h(str(batch.report_date))}</strong></div>
      </div>
    </header>
    <section>{body}</section>
  </article>
</body>
</html>"""


def build_akp_body(rows: list[DwdAgriAkpDensityDaily]) -> str:
    if not rows:
        return '<div class="empty">没有可展示的铲果密度 DWD 明细数据。</div>'

    detail_rows = [row for row in rows if row.row_label == "detail"]
    subtotal_rows = [row for row in rows if row.row_label in {"subtotal", "total"}]
    table_rows = "\n".join(
        f"""<tr class="{h(row.row_label)}">
          <td>{h(row.division)}</td>
          <td>{h(row.blok)}</td>
          <td>{h(row.sap)}</td>
          <td>{fmt(row.luas_ha)}</td>
          <td>{fmt(row.tt_year)}</td>
          <td>{fmt(row.panen_count)}</td>
          <td>{fmt(row.akp_percent)}%</td>
          <td>{fmt(row.panen_kg)}</td>
          <td>{fmt(row.jumlah_janjang)}</td>
          <td>{fmt(row.tk_panen)}</td>
          <td>{h(row.keterangan)}</td>
          <td>{h(row.row_label)}</td>
        </tr>"""
        for row in rows
    )
    return f"""
      <div class="summary">
        <div><span>明细行数</span><strong>{len(detail_rows)}</strong></div>
        <div><span>小计/合计行</span><strong>{len(subtotal_rows)}</strong></div>
        <div><span>总 Panen Kg</span><strong>{fmt(sum_number(row.panen_kg for row in detail_rows))}</strong></div>
        <div><span>总 Janjang</span><strong>{fmt(sum_number(row.jumlah_janjang for row in detail_rows))}</strong></div>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>小区</th><th>Blok</th><th>SAP</th><th>Luas Ha</th><th>TT</th>
              <th>Panen</th><th>AKP%</th><th>Panen Kg</th><th>Janjang</th>
              <th>TK Panen</th><th>Keterangan</th><th>行类型</th>
            </tr>
          </thead>
          <tbody>{table_rows}</tbody>
        </table>
      </div>"""


def build_attendance_body(rows: list[DwdAgriAttendanceDaily]) -> str:
    if not rows:
        return '<div class="empty">没有可展示的铲果工/养护工出勤 DWD 明细数据。</div>'

    detail_rows = [row for row in rows if row.row_label == "detail"]
    harvester_rows = [row for row in detail_rows if row.worker_type == "harvester"]
    maintenance_rows = [row for row in detail_rows if row.worker_type == "maintenance"]
    table_rows = "\n".join(
        f"""<tr class="{h(row.row_label)}">
          <td>{h(worker_type_label(row.worker_type))}</td>
          <td>{h(row.afdeling)}</td>
          <td>{fmt(row.luas_ha)}</td>
          <td>{fmt(row.kebutuhan_pemanen)}</td>
          <td>{fmt(row.actual_pemanen)}</td>
          <td>{fmt(row.actual_vs_kebutuhan)}</td>
          <td>{fmt(row.hadir)}</td>
          <td>{fmt(row.hadir_percent)}%</td>
          <td>{fmt(row.ijin)}</td>
          <td>{fmt(row.ijin_percent)}%</td>
          <td>{fmt(row.total_karyawan)}</td>
          <td>{fmt(row.total_percent)}%</td>
          <td>{h(row.row_label)}</td>
        </tr>"""
        for row in rows
    )
    return f"""
      <div class="summary">
        <div><span>铲果工实际人数</span><strong>{fmt(sum_number(row.actual_pemanen for row in harvester_rows))}</strong></div>
        <div><span>铲果工出勤人数</span><strong>{fmt(sum_number(row.hadir for row in harvester_rows))}</strong></div>
        <div><span>养护工实际人数</span><strong>{fmt(sum_number(row.actual_pemanen for row in maintenance_rows))}</strong></div>
        <div><span>养护工出勤人数</span><strong>{fmt(sum_number(row.hadir for row in maintenance_rows))}</strong></div>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>人员类型</th><th>Afdeling</th><th>Luas Ha</th><th>需求人数</th>
              <th>实际人数</th><th>差异</th><th>出勤人数</th><th>出勤率</th>
              <th>Ijin</th><th>Ijin率</th><th>总人数</th><th>总率</th><th>行类型</th>
            </tr>
          </thead>
          <tbody>{table_rows}</tbody>
        </table>
      </div>"""


def sum_number(values: Any) -> float:
    return sum(value for value in values if value is not None)


def sum_decimal(values: Any) -> float:
    total = 0.0
    for value in values:
        if value is None:
            continue
        try:
            total += float(value)
        except (TypeError, ValueError):
            continue
    return total


def average_percent(values: Any) -> str:
    numbers = [float(value) for value in values if value is not None]
    if not numbers:
        return ""
    return f"{sum(numbers) / len(numbers):.1f}%"


def safe_ratio(numerator: Any, denominator: Any) -> float | None:
    try:
        denominator_value = float(denominator)
        if denominator_value == 0:
            return None
        return float(numerator) / denominator_value
    except (TypeError, ValueError):
        return None


def fmt_percent(value: Any) -> str:
    if value is None:
        return ""
    try:
        return f"{float(value) * 100:.1f}%"
    except (TypeError, ValueError):
        return ""


def fmt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return f"{int(value):,}"
        return f"{value:,.2f}".rstrip("0").rstrip(".")
    if isinstance(value, int):
        return f"{value:,}"
    return h(str(value))


def percent_value(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value) / 100
    except (TypeError, ValueError):
        return None


def worker_type_label(worker_type: str) -> str:
    labels = {
        "harvester": "铲果工",
        "maintenance": "养护工",
        "unknown": "未识别",
    }
    return labels.get(worker_type, worker_type or "未识别")


def h(value: str) -> str:
    return escape(value, quote=True)
