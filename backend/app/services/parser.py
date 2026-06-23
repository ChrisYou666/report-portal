from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from typing import Any, Callable, Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models import (
    DwdAgriAkpDensityDaily,
    DwdAgriAttendanceDaily,
    DwdAgriHarvestDaily,
    DwdAgriProductionTargetMonthly,
    ParsedDocument,
    ParsedField,
    ParsedStructuredRecord,
    UploadBatch,
    UploadFileRecord,
)

# 向后兼容别名（旧代码引用）
DwdAgriAkpDensityDaily = DwdAgriAkpDensityDaily
DwdAgriAttendanceDaily = DwdAgriAttendanceDaily
DwdAgriHarvestDaily = DwdAgriHarvestDaily
DwdAgriProductionTargetMonthly = DwdAgriProductionTargetMonthly
DwdAgriProductionTargetMonthly = DwdAgriProductionTargetMonthly


IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "bmp", "tif", "tiff"}
EXCEL_EXTENSIONS = {"xlsx"}
CSV_EXTENSIONS = {"csv"}

_ocr_engine: Any | None = None


ParseProgressCallback = Callable[[dict[str, Any]], None]


def parse_batch(
    batch: UploadBatch,
    db: Session,
    progress_callback: Optional[ParseProgressCallback] = None,
) -> str:
    clear_existing_parse_results(batch, db)

    parsed_count = 0
    failed_count = 0
    skipped_count = 0
    total_files = len(batch.files)

    if progress_callback:
        progress_callback(
            {
                "status": "running",
                "total_files": total_files,
                "processed_files": 0,
                "parsed_files": 0,
                "skipped_files": 0,
                "failed_files": 0,
                "current_filename": "",
                "message": "开始解析批次。",
            }
        )

    for index, file_record in enumerate(batch.files, start=1):
        if progress_callback:
            progress_callback(
                {
                    "status": "running",
                    "total_files": total_files,
                    "processed_files": index - 1,
                    "parsed_files": parsed_count,
                    "skipped_files": skipped_count,
                    "failed_files": failed_count,
                    "current_filename": file_record.original_filename,
                    "message": f"正在解析：{file_record.original_filename}",
                }
            )

        document = parse_file(batch, file_record, db)
        if document.status == "parsed":
            parsed_count += 1
            file_record.status = "parsed"
        elif document.status == "parse_skipped":
            skipped_count += 1
            file_record.status = "parse_skipped"
        else:
            failed_count += 1
            file_record.status = "parse_failed"

        if progress_callback:
            progress_callback(
                {
                    "status": "running",
                    "total_files": total_files,
                    "processed_files": index,
                    "parsed_files": parsed_count,
                    "skipped_files": skipped_count,
                    "failed_files": failed_count,
                    "current_filename": file_record.original_filename,
                    "message": f"已完成 {index}/{total_files} 个文件。",
                }
            )

    if failed_count:
        batch.status = "parse_failed"
    elif parsed_count:
        batch.status = "parsed"
    else:
        batch.status = "parse_skipped"

    message = (
        f"批次 {batch.batch_no} 解析完成："
        f"成功 {parsed_count} 个文件，跳过 {skipped_count} 个文件，失败 {failed_count} 个文件。"
    )
    if progress_callback:
        progress_callback(
            {
                "status": batch.status,
                "total_files": total_files,
                "processed_files": total_files,
                "parsed_files": parsed_count,
                "skipped_files": skipped_count,
                "failed_files": failed_count,
                "current_filename": "",
                "message": message,
            }
        )

    return message


def clear_existing_parse_results(batch: UploadBatch, db: Session) -> None:
    db.query(DwdAgriAkpDensityDaily).filter(DwdAgriAkpDensityDaily.batch_id == batch.id).delete(synchronize_session=False)
    db.query(DwdAgriAttendanceDaily).filter(DwdAgriAttendanceDaily.batch_id == batch.id).delete(synchronize_session=False)
    db.query(DwdAgriHarvestDaily).filter(DwdAgriHarvestDaily.batch_id == batch.id).delete(synchronize_session=False)
    db.query(DwdAgriProductionTargetMonthly).filter(DwdAgriProductionTargetMonthly.batch_id == batch.id).delete(synchronize_session=False)
    db.query(DwdAgriProductionTargetMonthly).filter(DwdAgriProductionTargetMonthly.batch_id == batch.id).delete(synchronize_session=False)
    existing_documents = db.scalars(
        select(ParsedDocument).where(ParsedDocument.batch_id == batch.id)
    ).all()
    for document in existing_documents:
        db.delete(document)
    db.flush()


def parse_file(batch: UploadBatch, file_record: UploadFileRecord, db: Session) -> ParsedDocument:
    source_path = resolve_local_file_path(batch, file_record)
    parser_type = detect_parser_type(file_record)
    document = ParsedDocument(
        batch_id=batch.id,
        file_id=file_record.id,
        parser_type=parser_type,
        source_path=str(source_path),
        status="parsed",
    )
    db.add(document)
    db.flush()

    if not source_path.exists():
        document.status = "parse_failed"
        document.error_message = f"本地原文件不存在：{source_path}"
        return document

    try:
        if parser_type == "excel":
            parse_excel_file(document, source_path, db)
        elif parser_type == "csv":
            parse_csv_file(document, source_path, db)
        elif parser_type == "ocr":
            parse_image_file(document, source_path, db)
        else:
            document.status = "parse_skipped"
            document.error_message = f"暂不支持解析该文件类型：{file_record.file_type}"
        if document.status == "parsed":
            db.flush()
            load_dwd_records(document, db)
    except Exception as exc:  # Keep one bad file from killing the whole batch.
        document.status = "parse_failed"
        document.error_message = str(exc)

    return document


def resolve_local_file_path(batch: UploadBatch, file_record: UploadFileRecord) -> Path:
    stored_path = Path(file_record.stored_path)
    if stored_path.exists():
        return stored_path

    local_copy = Path(settings.storage_dir) / "originals" / batch.batch_no / Path(file_record.original_filename).name
    if local_copy.exists():
        return local_copy

    return local_copy


def detect_parser_type(file_record: UploadFileRecord) -> str:
    extension = file_record.file_type.lower().strip(".")
    if extension in EXCEL_EXTENSIONS:
        return "excel"
    if extension in CSV_EXTENSIONS:
        return "csv"
    if extension in IMAGE_EXTENSIONS:
        return "ocr"
    return "unsupported"


def parse_excel_file(document: ParsedDocument, path: Path, db: Session) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("解析 Excel 需要安装 openpyxl：pip install openpyxl") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    raw_rows: list[dict[str, Any]] = []
    raw_text_parts: list[str] = []

    for worksheet in workbook.worksheets:
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            for column_index, value in enumerate(row, start=1):
                if value is None or str(value).strip() == "":
                    continue
                text_value = str(value).strip()
                raw_rows.append(
                    {
                        "sheet": worksheet.title,
                        "row": row_index,
                        "column": column_index,
                        "value": text_value,
                    }
                )
                raw_text_parts.append(text_value)
                db.add(
                    ParsedField(
                        parsed_document_id=document.id,
                        batch_id=document.batch_id,
                        file_id=document.file_id,
                        record_type="excel_cell",
                        sheet_name=worksheet.title,
                        row_index=row_index,
                        column_index=column_index,
                        field_value=text_value,
                    )
                )

    document.raw_text = "\n".join(raw_text_parts)
    document.raw_json = json.dumps(raw_rows, ensure_ascii=False)


def parse_csv_file(document: ParsedDocument, path: Path, db: Session) -> None:
    rows = read_csv_rows(path)
    raw_cells: list[dict[str, Any]] = []
    raw_text_parts: list[str] = []

    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            text_value = value.strip()
            if not text_value:
                continue
            raw_cells.append({"row": row_index, "column": column_index, "value": text_value})
            raw_text_parts.append(text_value)
            db.add(
                ParsedField(
                    parsed_document_id=document.id,
                    batch_id=document.batch_id,
                    file_id=document.file_id,
                    record_type="csv_cell",
                    row_index=row_index,
                    column_index=column_index,
                    field_value=text_value,
                )
            )

    document.raw_text = "\n".join(raw_text_parts)
    document.raw_json = json.dumps(raw_cells, ensure_ascii=False)


def read_csv_rows(path: Path) -> list[list[str]]:
    for encoding in ["utf-8-sig", "utf-8", "gbk"]:
        try:
            with path.open("r", encoding=encoding, newline="") as file_obj:
                return list(csv.reader(file_obj))
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"CSV 编码无法识别：{path}")


def parse_image_file(document: ParsedDocument, path: Path, db: Session) -> None:
    ocr_lines = run_paddle_ocr(path)
    raw_text_parts: list[str] = []

    for row_index, line in enumerate(ocr_lines, start=1):
        text_value = line.get("text", "").strip()
        if not text_value:
            continue
        raw_text_parts.append(text_value)
        db.add(
            ParsedField(
                parsed_document_id=document.id,
                batch_id=document.batch_id,
                file_id=document.file_id,
                record_type="ocr_line",
                row_index=row_index,
                field_value=text_value,
                confidence=line.get("confidence"),
                raw_json=json.dumps(line, ensure_ascii=False),
            )
        )

    document.raw_text = "\n".join(raw_text_parts)
    document.raw_json = json.dumps(ocr_lines, ensure_ascii=False)
    structure_ocr_records(document, path, ocr_lines, db)


def run_paddle_ocr(path: Path) -> list[dict[str, Any]]:
    engine = get_ocr_engine()

    if hasattr(engine, "ocr"):
        try:
            result = engine.ocr(str(path), cls=True)
        except TypeError:
            result = engine.ocr(str(path))
    elif hasattr(engine, "predict"):
        try:
            result = engine.predict(input=str(path))
        except TypeError:
            result = engine.predict(str(path))
    else:
        raise RuntimeError("当前 PaddleOCR 对象不支持 ocr 或 predict 方法")

    return normalize_ocr_result(result)


def get_ocr_engine() -> Any:
    global _ocr_engine
    if _ocr_engine is not None:
        return _ocr_engine

    ocr_cache = Path(settings.paddle_ocr_cache_dir)
    ocr_cache.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("PADDLEOCR_HOME", str(ocr_cache))
    os.environ.setdefault("PADDLE_HOME", str(ocr_cache))
    os.environ.setdefault("PADDLE_PDX_CACHE_HOME", str(ocr_cache))
    os.environ.setdefault("PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK", "True")
    os.environ.setdefault("MODELSCOPE_CACHE", str(ocr_cache / "modelscope"))

    from paddleocr import PaddleOCR

    base_kwargs = {
        "lang": "ch",
        "use_doc_orientation_classify": False,
        "use_doc_unwarping": False,
        "use_textline_orientation": False,
    }
    if settings.paddle_ocr_det_model_dir:
        base_kwargs["text_detection_model_dir"] = settings.paddle_ocr_det_model_dir
    if settings.paddle_ocr_rec_model_dir:
        base_kwargs["text_recognition_model_dir"] = settings.paddle_ocr_rec_model_dir
    if settings.paddle_ocr_textline_model_dir:
        base_kwargs["textline_orientation_model_dir"] = settings.paddle_ocr_textline_model_dir
    init_attempts = [
        base_kwargs,
        {"lang": "ch", "use_textline_orientation": False},
        {"lang": "ch"},
    ]
    last_error: Exception | None = None
    for kwargs in init_attempts:
        try:
            _ocr_engine = PaddleOCR(**kwargs)
            return _ocr_engine
        except TypeError as exc:
            last_error = exc

    raise RuntimeError(f"PaddleOCR 初始化失败：{last_error}")


def normalize_ocr_result(result: Any) -> list[dict[str, Any]]:
    lines: list[dict[str, Any]] = []
    collect_ocr_lines(result, lines)
    return lines


def collect_ocr_lines(value: Any, lines: list[dict[str, Any]]) -> None:
    if value is None:
        return

    if isinstance(value, dict):
        rec_texts = value.get("rec_texts")
        rec_scores = value.get("rec_scores", [])
        rec_boxes = first_present(value.get("rec_boxes"), value.get("dt_polys"), [])
        if isinstance(rec_texts, list):
            for index, text in enumerate(rec_texts):
                lines.append(
                    {
                        "text": str(text),
                        "confidence": safe_float(rec_scores[index]) if index < len(rec_scores) else None,
                        "box": to_jsonable(rec_boxes[index]) if index < len(rec_boxes) else None,
                    }
                )
            return

        if "text" in value:
            lines.append(
                {
                    "text": str(value.get("text", "")),
                    "confidence": safe_float(first_present(value.get("confidence"), value.get("score"))),
                    "box": to_jsonable(value.get("box")),
                }
            )
            return

        for nested in value.values():
            collect_ocr_lines(nested, lines)
        return

    if isinstance(value, (list, tuple)):
        if is_v2_line(value):
            text, confidence = value[1]
            lines.append(
                {
                    "text": str(text),
                    "confidence": safe_float(confidence),
                    "box": to_jsonable(value[0]),
                }
            )
            return

        for nested in value:
            collect_ocr_lines(nested, lines)


def is_v2_line(value: Any) -> bool:
    return (
        isinstance(value, (list, tuple))
        and len(value) == 2
        and isinstance(value[1], (list, tuple))
        and len(value[1]) >= 2
        and isinstance(value[1][0], str)
    )


def first_present(*values: Any) -> Any:
    for value in values:
        if value is not None:
            return value
    return None


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_jsonable(value: Any) -> Any:
    if hasattr(value, "tolist"):
        return value.tolist()
    if isinstance(value, tuple):
        return [to_jsonable(item) for item in value]
    if isinstance(value, list):
        return [to_jsonable(item) for item in value]
    if isinstance(value, dict):
        return {str(key): to_jsonable(item) for key, item in value.items()}
    return value


def structure_ocr_records(
    document: ParsedDocument,
    path: Path,
    ocr_lines: list[dict[str, Any]],
    db: Session,
) -> None:
    items = [item for item in (normalize_ocr_item(line) for line in ocr_lines) if item]
    rows = group_items_by_row(items)
    template_name = detect_ocr_template(ocr_lines)

    for row_index, row in enumerate(rows, start=1):
        cells = [
            {
                "text": item["text"],
                "box": item["box"],
                "confidence": item["confidence"],
            }
            for item in row
        ]
        db.add(
            ParsedStructuredRecord(
                parsed_document_id=document.id,
                batch_id=document.batch_id,
                file_id=document.file_id,
                template_name=template_name,
                record_type="ocr_table_row",
                row_index=row_index,
                record_json=json.dumps(
                    {
                        "row_text": " ".join(item["text"] for item in row),
                        "cells": cells,
                    },
                    ensure_ascii=False,
                ),
                confidence=average_confidence(row),
            )
        )

    if template_name == "akp_density":
        add_akp_density_records(document, rows, db)
    elif template_name == "harvester_attendance":
        add_harvester_attendance_records(document, rows, db)
    elif template_name == "production_monitoring":
        add_production_monitoring_records(document, rows, db)
    elif template_name == "production_budget":
        add_production_budget_records(document, rows, db)
    elif template_name == "production_estimate":
        add_production_estimate_records(document, rows, db)


def normalize_ocr_item(line: dict[str, Any]) -> dict[str, Any] | None:
    text = str(line.get("text") or "").strip()
    box = line.get("box")
    if not text or not isinstance(box, list) or len(box) < 4:
        return None

    if all(isinstance(value, (int, float)) for value in box[:4]):
        x1, y1, x2, y2 = [float(value) for value in box[:4]]
    elif isinstance(box[0], list):
        xs = [float(point[0]) for point in box if isinstance(point, list) and len(point) >= 2]
        ys = [float(point[1]) for point in box if isinstance(point, list) and len(point) >= 2]
        if not xs or not ys:
            return None
        x1, x2 = min(xs), max(xs)
        y1, y2 = min(ys), max(ys)
    else:
        return None

    return {
        "text": text,
        "box": [x1, y1, x2, y2],
        "x_center": (x1 + x2) / 2,
        "y_center": (y1 + y2) / 2,
        "height": max(y2 - y1, 1),
        "confidence": safe_float(line.get("confidence")),
    }


def group_items_by_row(items: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    if not items:
        return []

    sorted_items = sorted(items, key=lambda item: (item["y_center"], item["x_center"]))
    row_threshold = max(10.0, median([item["height"] for item in sorted_items]) * 0.75)
    rows: list[list[dict[str, Any]]] = []

    for item in sorted_items:
        if not rows:
            rows.append([item])
            continue

        current_row = rows[-1]
        current_y = sum(row_item["y_center"] for row_item in current_row) / len(current_row)
        if abs(item["y_center"] - current_y) <= row_threshold:
            current_row.append(item)
        else:
            rows.append([item])

    return [sorted(row, key=lambda item: item["x_center"]) for row in rows]


def detect_ocr_template(ocr_lines: list[dict[str, Any]]) -> str:
    joined_text = normalize_template_text(" ".join(str(line.get("text") or "") for line in ocr_lines[:160]))
    if is_production_budget_template(joined_text):
        return "production_budget"
    if is_production_monitoring_template(joined_text):
        return "production_monitoring"
    if is_production_estimate_template(joined_text):
        return "production_estimate"
    if is_harvester_attendance_template(joined_text):
        return "harvester_attendance"
    if is_akp_density_template(joined_text):
        return "akp_density"
    return "generic_ocr_table"


def normalize_template_text(value: str) -> str:
    return value.upper().replace("％", "%")


def is_akp_density_template(text: str) -> bool:
    negative_tokens = ["ESTIMASI", "预计", "TAKSASI PRODUKSI", "PRODUKSI", "JANJANG DI KIRIM", "TRIP", "HM"]
    if any(token in text for token in negative_tokens):
        return False

    required_groups = [
        ["AKP"],
        ["BLOK"],
        ["LUAS"],
        ["JANJANG", "JJG"],
        ["KG", "PANEN KG"],
    ]
    matched_groups = sum(1 for group in required_groups if any(token in text for token in group))
    return matched_groups >= 4


def is_harvester_attendance_template(text: str) -> bool:
    if "KEHADIRAN" not in text and "HADIR" not in text:
        return False
    worker_signal = any(token in text for token in ["PEMANEN", "PERAWATAN", "PERAWAT"])
    attendance_signal = any(token in text for token in ["IJIN", "CUTI", "SAKIT", "MANGKIR", "ACTUAL", "KEBUTUHAN"])
    return worker_signal and attendance_signal


def is_production_monitoring_template(text: str) -> bool:
    compact_text = text.replace(" ", "")
    title_signal = (
        "PENCAPAIANPRODUKSI" in compact_text
        or ("MONITORING" in text and "PENCAPAIAN" in text and "PRODUKSI" in text)
    )
    column_signal = "DIVISI" in text and "BBC" in text and "TARGET" in text and "HARIAN" in text
    return title_signal and column_signal


def is_production_budget_template(text: str) -> bool:
    compact_text = text.replace(" ", "")
    title_signal = "PERENCANAANPRODUKSI" in compact_text or "PRODUKSI2025/2026" in compact_text
    column_signal = "BUDGET" in text and "SEP-25" in text and "MEI-26" in text and "YIELD" in text
    return title_signal and column_signal


def is_production_estimate_template(text: str) -> bool:
    compact_text = text.replace(" ", "")
    title_signal = "ESTIMASIPRODUKSI" in compact_text or "预计产量" in text
    attendance_signal = "KEHADIRAN" in text or "出勤" in text
    column_signal = "DIVISI" in text and "AKP" in text and ("PRODUKSI" in text or "产量" in text)
    return title_signal and attendance_signal and column_signal


def add_akp_density_records(document: ParsedDocument, rows: list[list[dict[str, Any]]], db: Session) -> None:
    columns = [
        ("division", 50),
        ("blok", 190),
        ("sap", 320),
        ("luas_ha", 495),
        ("tt", 585),
        ("panen_count", 690),
        ("akp_percent", 805),
        ("panen_kg", 965),
        ("jumlah_janjang", 1125),
        ("tk_panen", 1285),
        ("keterangan", 1375),
    ]
    current_division = ""
    pending_division_records: list[ParsedStructuredRecord] = []
    infer_next_division = False

    for row_index, row in enumerate(rows, start=1):
        row_text = " ".join(item["text"] for item in row).upper()
        if is_header_or_title(row_text):
            continue

        mapped = map_row_to_columns(row, columns)
        if len(row) == 1 and row[0]["x_center"] < 120 and row[0]["text"].strip().isalpha():
            current_division = row[0]["text"].strip()
            continue

        has_business_value = any(mapped.get(key) for key in ["blok", "sap", "luas_ha", "panen_count", "panen_kg"])
        is_total_row = any(token in row_text for token in ["SUB TOTAL", "TOTAL"])
        if not has_business_value and not is_total_row:
            continue

        mapped_division = mapped.get("division", "").strip()
        if mapped_division and mapped_division.isalpha() and len(mapped_division) <= 3:
            current_division = mapped_division
            infer_next_division = False
            for record in pending_division_records:
                set_structured_record_data_value(record, "division", current_division)
            pending_division_records = []
        elif infer_next_division and current_division == "C" and looks_like_d_division_akp_row(mapped):
            current_division = "D"
            mapped["division"] = current_division
            infer_next_division = False
        elif current_division:
            mapped["division"] = current_division
        mapped["row_label"] = "subtotal" if "SUB TOTAL" in row_text else "total" if "TOTAL" in row_text else "detail"
        record = add_structured_record(document, "akp_density", "akp_density_row", row_index, mapped, row, db)
        if not mapped.get("division"):
            pending_division_records.append(record)
        if mapped["row_label"] == "subtotal" and current_division == "C":
            infer_next_division = True


def add_harvester_attendance_records(document: ParsedDocument, rows: list[list[dict[str, Any]]], db: Session) -> None:
    columns = [
        ("afdeling", 50),
        ("luas_ha", 130),
        ("kebutuhan_pemanen", 225),
        ("actual_pemanen", 322),
        ("actual_vs_kebutuhan", 421),
        ("hadir", 503),
        ("hadir_percent", 570),
        ("ijin", 640),
        ("ijin_percent", 705),
        ("cuti", 770),
        ("cuti_percent", 835),
        ("sakit", 900),
        ("sakit_percent", 965),
        ("mangkir", 1035),
        ("mangkir_percent", 1100),
        ("total_karyawan", 1175),
        ("total_percent", 1245),
    ]

    section = ""
    worker_type = ""
    for row_index, row in enumerate(rows, start=1):
        row_text = " ".join(item["text"] for item in row).upper()
        if "MONITORING" in row_text:
            section = "daily_attendance"
            worker_type = detect_worker_type(row_text, worker_type)
            continue
        if "PANTAUAN" in row_text or "JAMKERJA" in row_text:
            section = "work_hour_monitoring"
            worker_type = detect_worker_type(row_text, worker_type)
            continue
        if "KEHADIRAN" in row_text:
            worker_type = detect_worker_type(row_text, worker_type)
            continue
        if is_header_or_title(row_text):
            continue

        mapped = map_row_to_columns(row, columns)
        afdeling = mapped.get("afdeling", "")
        if not afdeling or len(afdeling) > 12:
            continue

        has_numbers = sum(1 for key, value in mapped.items() if key != "afdeling" and value)
        if has_numbers < 2:
            continue

        mapped["section"] = section or "attendance"
        mapped["worker_type"] = worker_type or "unknown"
        mapped["row_label"] = "total" if afdeling.upper() == "TOTAL" else "detail"
        add_structured_record(document, "harvester_attendance", "harvester_attendance_row", row_index, mapped, row, db)


def add_production_monitoring_records(document: ParsedDocument, rows: list[list[dict[str, Any]]], db: Session) -> None:
    columns = [
        ("division", 80),
        ("luas_ha", 175),
        ("bbc_ton", 265),
        ("actual_today_ton", 365),
        ("actual_to_date_ton", 465),
        ("actual_vs_bbc_percent", 545),
        ("remaining_bbc_ton", 625),
        ("remaining_effective_days", 700),
        ("daily_target_ton", 765),
    ]

    for row_index, row in enumerate(rows, start=1):
        row_text = " ".join(item["text"] for item in row).upper()
        if is_header_or_title(row_text):
            continue

        mapped = map_row_to_columns(row, columns)
        division = normalize_division_value(mapped.get("division", ""))
        if division not in {"A", "B", "C", "D", "TOTAL"}:
            continue

        has_numbers = sum(1 for key, value in mapped.items() if key != "division" and value)
        if has_numbers < 3:
            continue

        mapped["division"] = division
        mapped["row_label"] = "total" if division == "TOTAL" else "detail"
        add_structured_record(
            document,
            "production_monitoring",
            "production_monitoring_row",
            row_index,
            mapped,
            row,
            db,
        )


def add_production_budget_records(document: ParsedDocument, rows: list[list[dict[str, Any]]], db: Session) -> None:
    columns = [
        ("division", 200),
        ("mature_area_ha", 315),
        ("budget_sep_ton", 415),
        ("budget_oct_ton", 505),
        ("budget_nov_ton", 595),
        ("budget_dec_ton", 685),
        ("budget_jan_ton", 774),
        ("budget_feb_ton", 863),
        ("budget_mar_ton", 953),
        ("budget_apr_ton", 1032),
        ("budget_may_ton", 1132),
        ("budget_jun_ton", 1222),
        ("budget_jul_ton", 1311),
        ("budget_aug_ton", 1401),
        ("annual_budget_ton", 1497),
        ("yield_ton_per_ha", 1572),
    ]

    for row_index, row in enumerate(rows, start=1):
        row_text = " ".join(item["text"] for item in row).upper()
        if "BUDGET" in row_text or "YIELD" in row_text or "SEP-25" in row_text:
            continue

        mapped = map_row_to_columns(row, columns)
        division = normalize_budget_division_value(mapped.get("division", ""))
        if division not in {"A", "B", "C", "D", "TOTAL"}:
            continue

        has_numbers = sum(1 for key, value in mapped.items() if key != "division" and value)
        if has_numbers < 6:
            continue

        mapped["division"] = division
        mapped["row_label"] = "total" if division == "TOTAL" else "detail"
        add_structured_record(
            document,
            "production_budget",
            "production_budget_row",
            row_index,
            mapped,
            row,
            db,
        )


def add_production_estimate_records(document: ParsedDocument, rows: list[list[dict[str, Any]]], db: Session) -> None:
    columns = [
        ("row_no", 120),
        ("division", 225),
        ("mature_area_ha", 523),
        ("estimated_harvest_area_ha", 657),
        ("estimated_production_kg", 888),
        ("akp_percent", 1011),
    ]

    seen_data_row = False
    for row_index, row in enumerate(rows, start=1):
        row_text = " ".join(item["text"] for item in row).upper()
        if "KEHADIRAN" in row_text:
            if seen_data_row:
                break
            continue
        if any(token in row_text for token in ["LUAS TM", "预计铲果面积", "ESTIMASI"]) and "TOTAL" not in row_text:
            continue

        mapped = map_row_to_columns(row, columns)
        division = normalize_budget_division_value(mapped.get("division", ""))
        if division not in {"A", "B", "C", "D", "TOTAL"}:
            continue

        has_numbers = sum(
            1
            for key in ["mature_area_ha", "estimated_harvest_area_ha", "estimated_production_kg", "akp_percent"]
            if mapped.get(key)
        )
        if has_numbers < 3:
            continue

        mapped["division"] = division
        mapped["row_label"] = "total" if division == "TOTAL" else "detail"
        seen_data_row = True
        add_structured_record(
            document,
            "production_estimate",
            "production_estimate_row",
            row_index,
            mapped,
            row,
            db,
        )


def detect_worker_type(row_text: str, current_worker_type: str = "") -> str:
    if "PERAWATAN" in row_text or "PERAWAT" in row_text:
        return "maintenance"
    if "PEMANEN" in row_text or "PANEN" in row_text:
        return "harvester"
    return current_worker_type


def looks_like_d_division_akp_row(mapped: dict[str, str]) -> bool:
    block = mapped.get("blok", "").strip().upper()
    sap = mapped.get("sap", "").strip()
    if not block or "TOTAL" in block:
        return False
    return block.startswith("B2") or sap in {"253", "327", "328", "329"}


def add_structured_record(
    document: ParsedDocument,
    template_name: str,
    record_type: str,
    row_index: int,
    data: dict[str, str],
    source_row: list[dict[str, Any]],
    db: Session,
) -> ParsedStructuredRecord:
    record = ParsedStructuredRecord(
        parsed_document_id=document.id,
        batch_id=document.batch_id,
        file_id=document.file_id,
        template_name=template_name,
        record_type=record_type,
        row_index=row_index,
        record_json=json.dumps(
            {
                "data": data,
                "source_cells": [
                    {
                        "text": item["text"],
                        "box": item["box"],
                        "confidence": item["confidence"],
                    }
                    for item in source_row
                ],
            },
            ensure_ascii=False,
        ),
        confidence=average_confidence(source_row),
    )
    db.add(record)
    return record


def set_structured_record_data_value(record: ParsedStructuredRecord, key: str, value: str) -> None:
    payload = json.loads(record.record_json)
    payload.setdefault("data", {})[key] = value
    record.record_json = json.dumps(payload, ensure_ascii=False)


def map_row_to_columns(row: list[dict[str, Any]], columns: list[tuple[str, int]]) -> dict[str, str]:
    mapped: dict[str, list[str]] = {}
    for item in row:
        field_name = nearest_column(item["x_center"], columns)
        mapped.setdefault(field_name, []).append(item["text"])
    return {field_name: " ".join(values).strip() for field_name, values in mapped.items()}


def nearest_column(x_center: float, columns: list[tuple[str, int]]) -> str:
    return min(columns, key=lambda column: abs(x_center - column[1]))[0]


def is_header_or_title(row_text: str) -> bool:
    header_tokens = [
        "BLOK",
        "PANEN",
        "TAKSASI",
        "JUMLAH",
        "DIVISI",
        "AFDELING",
        "PEMANEN",
        "KEBUTUHAN",
        "KEHADIRAN",
        "TIDAK HADIR",
        "HARI/TANGGAL",
        "ESTATE",
        "PT.",
    ]
    return any(token in row_text for token in header_tokens)


def average_confidence(row: list[dict[str, Any]]) -> float | None:
    values = [item["confidence"] for item in row if item.get("confidence") is not None]
    if not values:
        return None
    return sum(values) / len(values)


def median(values: list[float]) -> float:
    if not values:
        return 0.0
    sorted_values = sorted(values)
    middle = len(sorted_values) // 2
    if len(sorted_values) % 2:
        return sorted_values[middle]
    return (sorted_values[middle - 1] + sorted_values[middle]) / 2


def upsert_dwd_record(db: Session, model_class: Any, filters: dict, values: dict) -> None:
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    stmt = pg_insert(model_class).values(**filters, **values)
    stmt = stmt.on_conflict_do_update(
        index_elements=list(filters.keys()),
        set_={col: stmt.excluded[col] for col in values},
    )
    db.execute(stmt)


def load_dwd_records(document: ParsedDocument, db: Session) -> None:
    structured_records = db.scalars(
        select(ParsedStructuredRecord)
        .where(ParsedStructuredRecord.parsed_document_id == document.id)
        .order_by(ParsedStructuredRecord.row_index)
    ).all()

    for record in structured_records:
        if record.record_type == "akp_density_row":
            load_dwd_akp_density_record(document, record, db)
        elif record.record_type == "harvester_attendance_row":
            load_dwd_harvester_attendance_record(document, record, db)
        elif record.record_type == "production_monitoring_row":
            load_dwd_production_monitoring_record(document, record, db)
        elif record.record_type == "production_budget_row":
            load_dwd_production_budget_record(document, record, db)
        elif record.record_type == "production_estimate_row":
            load_dwd_production_estimate_record(document, record, db)


def load_dwd_akp_density_record(
    document: ParsedDocument,
    record: ParsedStructuredRecord,
    db: Session,
) -> None:
    data = get_record_data(record)
    quality_status, quality_message = validate_required_fields(data, ["division", "blok"])
    batch = document.batch
    row_label = data.get("row_label", "detail")

    upsert_dwd_record(
        db,
        DwdAgriAkpDensityDaily,
        filters={
            "report_date": batch.report_date,
            "site": batch.site,
            "department": batch.department,
            "division": data.get("division", ""),
            "blok": data.get("blok", ""),
            "row_label": row_label,
        },
        values={
            "batch_id": document.batch_id,
            "batch_no": batch.batch_no,
            "file_id": document.file_id,
            "source_record_id": record.id,
            "sap": data.get("sap", ""),
            "luas_ha": parse_number(data.get("luas_ha")),
            "tt_year": parse_int(data.get("tt")),
            "panen_count": parse_number(data.get("panen_count")),
            "akp_percent": parse_percent(data.get("akp_percent")),
            "panen_kg": parse_number(data.get("panen_kg")),
            "jumlah_janjang": parse_number(data.get("jumlah_janjang")),
            "tk_panen": parse_number(data.get("tk_panen")),
            "keterangan": data.get("keterangan", ""),
            "confidence": record.confidence,
            "quality_status": quality_status,
            "quality_message": quality_message,
        },
    )


def load_dwd_harvester_attendance_record(
    document: ParsedDocument,
    record: ParsedStructuredRecord,
    db: Session,
) -> None:
    data = get_record_data(record)
    quality_status, quality_message = validate_required_fields(data, ["section", "afdeling"])
    batch = document.batch
    row_label = data.get("row_label", "detail")

    upsert_dwd_record(
        db,
        DwdAgriAttendanceDaily,
        filters={
            "report_date": batch.report_date,
            "site": batch.site,
            "department": batch.department,
            "section": data.get("section", ""),
            "worker_type": data.get("worker_type", ""),
            "afdeling": data.get("afdeling", ""),
            "row_label": row_label,
        },
        values={
            "batch_id": document.batch_id,
            "batch_no": batch.batch_no,
            "file_id": document.file_id,
            "source_record_id": record.id,
            "luas_ha": parse_number(data.get("luas_ha")),
            "kebutuhan_pemanen": parse_number(data.get("kebutuhan_pemanen")),
            "actual_pemanen": parse_number(data.get("actual_pemanen")),
            "actual_vs_kebutuhan": parse_number(data.get("actual_vs_kebutuhan")),
            "hadir": parse_number(data.get("hadir")),
            "hadir_percent": parse_percent(data.get("hadir_percent")),
            "ijin": parse_number(data.get("ijin")),
            "ijin_percent": parse_percent(data.get("ijin_percent")),
            "cuti": parse_number(data.get("cuti")),
            "cuti_percent": parse_percent(data.get("cuti_percent")),
            "sakit": parse_number(data.get("sakit")),
            "sakit_percent": parse_percent(data.get("sakit_percent")),
            "mangkir": parse_number(data.get("mangkir")),
            "mangkir_percent": parse_percent(data.get("mangkir_percent")),
            "total_karyawan": parse_number(data.get("total_karyawan")),
            "total_percent": parse_percent(data.get("total_percent")),
            "confidence": record.confidence,
            "quality_status": quality_status,
            "quality_message": quality_message,
        },
    )


def load_dwd_production_monitoring_record(
    document: ParsedDocument,
    record: ParsedStructuredRecord,
    db: Session,
) -> None:
    data = get_record_data(record)
    quality_status, quality_message = validate_required_fields(data, ["division", "bbc_ton"])
    batch = document.batch
    row_label = data.get("row_label", "detail")

    upsert_dwd_record(
        db,
        DwdAgriHarvestDaily,
        filters={
            "report_date": batch.report_date,
            "site": batch.site,
            "department": batch.department,
            "division": data.get("division", ""),
            "row_label": row_label,
        },
        values={
            "batch_id": document.batch_id,
            "batch_no": batch.batch_no,
            "file_id": document.file_id,
            "source_record_id": record.id,
            "luas_ha": parse_localized_number(data.get("luas_ha")),
            "bbc_ton": parse_localized_number(data.get("bbc_ton")),
            "actual_today_ton": parse_localized_number(data.get("actual_today_ton")),
            "actual_to_date_ton": parse_localized_number(data.get("actual_to_date_ton")),
            "actual_vs_bbc_percent": parse_localized_number(data.get("actual_vs_bbc_percent")),
            "remaining_bbc_ton": parse_localized_number(data.get("remaining_bbc_ton")),
            "remaining_effective_days": parse_localized_number(data.get("remaining_effective_days")),
            "daily_target_ton": parse_localized_number(data.get("daily_target_ton")),
            "confidence": record.confidence,
            "quality_status": quality_status,
            "quality_message": quality_message,
        },
    )


def load_dwd_production_budget_record(
    document: ParsedDocument,
    record: ParsedStructuredRecord,
    db: Session,
) -> None:
    data = get_record_data(record)
    quality_status, quality_message = validate_required_fields(data, ["division", "mature_area_ha"])
    batch = document.batch
    row_label = data.get("row_label", "detail")

    upsert_dwd_record(
        db,
        DwdAgriProductionTargetMonthly,
        filters={
            "report_date": batch.report_date,
            "site": batch.site,
            "department": batch.department,
            "division": data.get("division", ""),
            "row_label": row_label,
        },
        values={
            "batch_id": document.batch_id,
            "batch_no": batch.batch_no,
            "file_id": document.file_id,
            "source_record_id": record.id,
            "mature_area_ha": parse_localized_number(data.get("mature_area_ha")),
            "budget_sep_ton": parse_localized_number(data.get("budget_sep_ton")),
            "budget_oct_ton": parse_localized_number(data.get("budget_oct_ton")),
            "budget_nov_ton": parse_localized_number(data.get("budget_nov_ton")),
            "budget_dec_ton": parse_localized_number(data.get("budget_dec_ton")),
            "budget_jan_ton": parse_localized_number(data.get("budget_jan_ton")),
            "budget_feb_ton": parse_localized_number(data.get("budget_feb_ton")),
            "budget_mar_ton": parse_localized_number(data.get("budget_mar_ton")),
            "budget_apr_ton": parse_localized_number(data.get("budget_apr_ton")),
            "budget_may_ton": parse_localized_number(data.get("budget_may_ton")),
            "budget_jun_ton": parse_localized_number(data.get("budget_jun_ton")),
            "budget_jul_ton": parse_localized_number(data.get("budget_jul_ton")),
            "budget_aug_ton": parse_localized_number(data.get("budget_aug_ton")),
            "annual_budget_ton": parse_localized_number(data.get("annual_budget_ton")),
            "yield_ton_per_ha": parse_localized_number(data.get("yield_ton_per_ha")),
            "confidence": record.confidence,
            "quality_status": quality_status,
            "quality_message": quality_message,
        },
    )


def load_dwd_production_estimate_record(
    document: ParsedDocument,
    record: ParsedStructuredRecord,
    db: Session,
) -> None:
    data = get_record_data(record)
    quality_status, quality_message = validate_required_fields(data, ["division", "estimated_production_kg"])
    batch = document.batch
    row_label = data.get("row_label", "detail")

    upsert_dwd_record(
        db,
        DwdAgriProductionTargetMonthly,
        filters={
            "report_date": batch.report_date,
            "site": batch.site,
            "department": batch.department,
            "division": data.get("division", ""),
            "row_label": row_label,
        },
        values={
            "batch_id": document.batch_id,
            "batch_no": batch.batch_no,
            "file_id": document.file_id,
            "source_record_id": record.id,
            "mature_area_ha": parse_localized_number(data.get("mature_area_ha")),
            "estimated_harvest_area_ha": parse_localized_number(data.get("estimated_harvest_area_ha")),
            "estimated_production_kg": parse_localized_number(data.get("estimated_production_kg")),
            "akp_percent": parse_percent(data.get("akp_percent")),
            "confidence": record.confidence,
            "quality_status": quality_status,
            "quality_message": quality_message,
        },
    )


def get_record_data(record: ParsedStructuredRecord) -> dict[str, str]:
    try:
        payload = json.loads(record.record_json)
    except ValueError:
        return {}
    data = payload.get("data", {})
    if not isinstance(data, dict):
        return {}
    return {str(key): "" if value is None else str(value).strip() for key, value in data.items()}


def normalize_division_value(value: str) -> str:
    normalized = value.strip().upper().replace(" ", "")
    if "TOTAL" in normalized:
        return "TOTAL"
    for token in ["A", "B", "C", "D"]:
        if normalized == token:
            return token
    return normalized


def normalize_budget_division_value(value: str) -> str:
    normalized = value.strip().upper().replace(" ", "")
    if "TOTAL" in normalized:
        return "TOTAL"
    normalized = normalized.replace("DIVISI", "").replace("区", "")
    for token in ["A", "B", "C", "D"]:
        if normalized == token or normalized.endswith(token):
            return token
    return normalized


def validate_required_fields(data: dict[str, str], required_fields: list[str]) -> tuple[str, str]:
    missing_fields = [field for field in required_fields if not data.get(field)]
    if missing_fields:
        return "need_review", f"缺少必填字段：{', '.join(missing_fields)}"
    return "ok", ""


def parse_number(value: str | None) -> float | None:
    if value is None:
        return None
    cleaned = value.strip().replace(",", "").replace("%", "")
    if cleaned in {"", "-", "—"}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_localized_number(value: str | None) -> float | None:
    if value is None:
        return None
    cleaned = (
        value.strip()
        .replace(" ", "")
        .replace("%", "")
        .replace("，", ",")
        .replace("。", ".")
    )
    if cleaned in {"", "-", "鈥?"}:
        return None
    cleaned = cleaned.replace("−", "-")

    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        integer_part, decimal_part = cleaned.rsplit(",", 1)
        if len(decimal_part) <= 2:
            cleaned = f"{integer_part}.{decimal_part}"
        else:
            cleaned = cleaned.replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_int(value: str | None) -> int | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return int(parsed)


def parse_percent(value: str | None) -> float | None:
    return parse_number(value)
